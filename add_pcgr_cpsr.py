#!/usr/bin/env python3
# This script will add the pcgr and cpsr .tsv output to the existing .xlsx file

import csv
import sys
import openpyxl
import os
import pandas as pd

# Original xlsx file
input_xlsx_file = sys.argv[1]
csv_file = sys.argv[2]
pcgr_tsv_file = sys.argv[3]
cpsr_tsv_file = sys.argv[4]

wb = openpyxl.load_workbook(input_xlsx_file)
if os.path.getsize(csv_file) != 0:
	worksheet = wb.create_sheet('append_final_concat')
	read_csv = csv.reader(open(csv_file, 'r', encoding='utf-8'), delimiter=',')
	for rows in read_csv:
		worksheet.append(rows)

if os.path.getsize(pcgr_tsv_file) != 0:
	worksheet = wb.create_sheet('pcgr')
	read_tsv1 = csv.reader(open(pcgr_tsv_file, 'r', encoding='utf-8'), delimiter='\t')
	for rows in read_tsv1:
		worksheet.append(rows)

if os.path.getsize(cpsr_tsv_file) != 0:
	worksheet = wb.create_sheet('cpsr')
	read_tsv2 = csv.reader(open(cpsr_tsv_file, 'r', encoding='utf-8'), delimiter='\t')
	for row2 in read_tsv2:
		worksheet.append(row2)

wb.save(input_xlsx_file)


xl_file = pd.read_excel(input_xlsx_file, sheet_name=None)
with pd.ExcelWriter('output_temp.xlsx') as writer:
	for sheet_names in xl_file.keys():
		df = pd.read_excel(input_xlsx_file, sheet_name=sheet_names)
		df.fillna(value= -1 , inplace=True)

		if 'VAF' in df:
			df["VAF"] = df["VAF"].astype(str).str.rstrip("%").astype(float)

		df = df.rename(columns={"VAF":"VAF (%)"})
		if 'PopFreqMax' in df:
			df["PopFreqMax"] = df["PopFreqMax"].astype(str).astype(float)

		if 'ALT_COUNT' in df:
			df["ALT_COUNT"] = df["ALT_COUNT"].astype(str).astype(int)

		df.to_excel(writer, sheet_name=sheet_names, index=False)
